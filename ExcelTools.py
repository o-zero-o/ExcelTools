import pandas as pd
from openpyxl import load_workbook

def get_data(file_path, sheet_name="Sheet1"):
    """
    用于提取 Excel 的数据，输出为字典
    :param sheet_name: 要提取的sheet页
    :param file_path: 要提取的文件路径
    :return: 输出字典数据
    """
    file = pd.read_excel(file_path, sheet_name=sheet_name)
    data = file.to_dict(orient='records')
    return data

def write_data(data, filename, sheet_name="Sheet1"):
    """
    用于写入字典列表的数据，注意这个方法会删除其他的sheet页
    :param data: 字典列表，格式如 [{},{}]
    :param filename: 输出的文件路径
    :param sheet_name: 要写入的sheet页
    :return:
    """
    df = pd.DataFrame(data)
    df.to_excel(filename, index=False, sheet_name=sheet_name)
    print(f"数据已写入 {filename}")


def write_data_add(data, filename, sheet_name="Sheet1"):
    """
    用于追加字典列表的数据，到原有的表中，注意 这个方法无法对已存在的sheet页进行修改
    :param data: 字典列表，格式如 [{},{}]
    :param filename: 输出的文件路径
    :param sheet_name: 要写入的sheet页
    :return:
    """
    # 创建一个 Excel writer 对象
    with pd.ExcelWriter(filename, mode='a', engine='openpyxl') as writer:
        df = pd.DataFrame(data)
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    print(f"数据已写入 {filename} 的 {sheet_name} 页")


def writer_cover(data, filename, sheet_name):
    """
    没做好
    :param data:
    :param filename:
    :param sheet_name:
    :return:
    """
    # 打开 Excel 文件
    # with pd.ExcelWriter(filename, engine='openpyxl') as excel_writer:
    excel_writer = pd.ExcelWriter(filename, engine='openpyxl')
    book = load_workbook(filename)
    excel_writer.book = book

    # 列出所有 sheet 页，以确定要删除的 sheet 是否存在
    sheet_to_delete = sheet_name
    if sheet_to_delete in book.sheetnames:
        sheet = book[sheet_name]
        book.remove(sheet)  # 删除指定的 sheet 页
    df = pd.DataFrame(data)
    df.to_excel(excel_writer, index=False, sheet_name=sheet_name)
    excel_writer.save()  # 保存文件
    excel_writer.close()  # 关闭 ExcelWriter

    print(f"数据覆写入 {filename} 的 '{sheet_to_delete}' 页。")

def reorder(filename, first_list):
    """
    用于将一个表中的指定的sheet 页移到前面
    :param filename:   文件路径
    :param first_list:  要放到前面的sheet页，注意排好序 ["sheet1","sheet2","sheet2"]
    :return:
    """
    excel_file = pd.ExcelFile(filename)
    sheet_names = excel_file.sheet_names

    # 重新排列 sheet 页的顺序
    new_sheet_order = first_list + [sheet for sheet in sheet_names if sheet not in first_list]

    # 将重新排列后的 sheet 页写回到新的 Excel 文件
    with pd.ExcelWriter(filename) as writer:
        for sheet_name in new_sheet_order:
            df = pd.read_excel(filename, sheet_name=sheet_name)  # 读取原始 sheet 页数据
            df.to_excel(writer, sheet_name=sheet_name, index=False)  # 写入到新文件中
    print("Sheet页顺序整理完成")
